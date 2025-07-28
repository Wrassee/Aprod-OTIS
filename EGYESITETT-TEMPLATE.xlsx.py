#!/usr/bin/env python3
"""
Unified OTIS Template Creator
Combines protocol questions and measurement questions into a single Excel template
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_unified_template():
    """Create a unified template with both protocol and measurement questions"""
    
    # Create workbook with two sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create named sheets
    wb.remove(wb.active)
    questions_sheet = wb.create_sheet("questions")
    protocol_sheet = wb.create_sheet("protocol") 
    
    # === QUESTIONS SHEET ===
    questions_data = [
        # Protocol Questions (from original system)
        ["question_id", "title_hu", "title_de", "type", "cell_reference", "unit", "calculation_formula", "calculation_inputs", "min_value", "max_value", "group_name", "group_order", "required", "placeholder"],
        
        # Általános adatok (General Data)
        ["1", "Átvevő neve", "Name des Abnehmers", "text", "F9", "", "", "", "", "", "Általános adatok", "1", "true", "Teljes név"],
        ["2", "Szerelő neve", "Name des Monteurs", "text", "Q9", "", "", "", "", "", "Általános adatok", "2", "true", "Teljes név"],
        ["3", "Irányítószám", "Postleitzahl", "number", "G13", "", "", "", "", "", "Általános adatok", "3", "true", "1234"],
        ["4", "Város", "Stadt", "text", "O13", "", "", "", "", "", "Általános adatok", "4", "true", "Városnév"],
        ["5", "Utca", "Straße", "text", "G14", "", "", "", "", "", "Általános adatok", "5", "true", "Utcanév"],
        ["6", "Házszám", "Hausnummer", "number", "N14", "", "", "", "", "", "Általános adatok", "6", "true", "1"],
        ["7", "Otis Lift-azonosító", "OTIS Aufzug-ID", "text", "O16", "", "", "", "", "", "Általános adatok", "7", "true", "ABC123"],
        ["8", "Projekt-azonosító", "Projekt-ID", "text", "O17", "", "", "", "", "", "Általános adatok", "8", "true", "PRJ123"],
        ["9", "Kirendeltség", "Außenstelle", "text", "O19", "", "", "", "", "", "Általános adatok", "9", "true", "Kirendeltség név"],
        
        # Gépház (Machine Room)
        ["10", "X", "X", "yes_no_na", "A68,B68,C68", "", "", "", "", "", "Gépház", "1", "true", ""],
        ["11", "Gépház", "Maschinenraum", "yes_no_na", "A75;A76;A77,B75;B76;B77,C75;C76;C77", "", "", "", "", "", "Gépház", "2", "true", ""],
        
        # Modernizációban érintett (Modernization Affected) 
        ["12", "Kérdések", "Fragen", "true_false", "Q25", "", "", "", "", "", "Modernizációban érintett", "1", "true", ""],
        ["13", "Kérdések", "Fragen", "true_false", "Q26", "", "", "", "", "", "Modernizációban érintett", "2", "true", ""],
        ["14", "Kérdések", "Fragen", "true_false", "Q27", "", "", "", "", "", "Modernizációban érintett", "3", "true", ""],
        ["15", "Kérdések", "Fragen", "true_false", "Q28", "", "", "", "", "", "Modernizációban érintett", "4", "true", ""],
        ["16", "Kérdések", "Fragen", "true_false", "Q29", "", "", "", "", "", "Modernizációban érintett", "5", "true", ""],
        ["17", "Kérdések", "Fragen", "true_false", "Q30", "", "", "", "", "", "Modernizációban érintett", "6", "true", ""],
        ["18", "Kérdések", "Fragen", "true_false", "Q31", "", "", "", "", "", "Modernizációban érintett", "7", "true", ""],
        ["19", "Kérdések", "Fragen", "true_false", "Q32", "", "", "", "", "", "Modernizációban érintett", "8", "true", ""],
        ["20", "Kérdések", "Fragen", "true_false", "Q33", "", "", "", "", "", "Modernizációban érintett", "9", "true", ""],
        ["21", "Kérdések", "Fragen", "true_false", "Q34", "", "", "", "", "", "Modernizációban érintett", "10", "true", ""],
        
        # MEASUREMENT DATA BLOCK - NEW UNIFIED SECTION
        ["m1", "Távolság kabintető és Aknatető között", "Abstand zwischen Kabinendach und Schachtkopf", "measurement", "I278", "mm", "", "", "", "", "Mérési adatok", "1", "true", "Mérés mm-ben"],
        ["m2", "Távolság kabintető legmagasabb pontja és Aknatető között", "Oberer Teil des Bogengangs am tiefsten Punkt vom Schachtkopf", "measurement", "N278", "mm", "", "", "", "", "Mérési adatok", "2", "true", "Mérés mm-ben"],
        ["m3", "Távolság az akna padló és az ellensúly puffer között", "Nachlaufweg bis das Gegengewicht den Puffer berührt", "measurement", "I280", "mm", "", "", "", "", "Mérési adatok", "3", "true", "Mérés mm-ben"],
        ["m4", "Effektív távolság A", "Effektiver Sicherheitsabstand A", "calculated", "I283", "mm", "m1 - m3", "m1,m3", "700", "9000", "Mérési adatok", "4", "true", "Automatikusan számolt"],
        ["m5", "Effektív távolság B", "Effektiver Sicherheitsabstand B", "calculated", "N283", "mm", "m2 - m3", "m2,m3", "400", "9000", "Mérési adatok", "5", "true", "Automatikusan számolt"],
    ]
    
    # Write questions data
    for row_idx, row_data in enumerate(questions_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            questions_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Style questions sheet header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col in range(1, len(questions_data[0]) + 1):
        cell = questions_sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust columns
    for column_cells in questions_sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        questions_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    # === PROTOCOL SHEET === 
    # Load existing protocol template as base
    try:
        # Copy the existing protocol template
        existing_protocol_path = "/home/runner/workspace/uploads/1753639478281-Abnahmeprotokoll Leer DE.xlsx"
        existing_wb = openpyxl.load_workbook(existing_protocol_path)
        existing_sheet = existing_wb.active
        
        # Copy all data from existing protocol
        for row in existing_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    new_cell = protocol_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font
                        new_cell.border = cell.border
                        new_cell.fill = cell.fill
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection
                        new_cell.alignment = cell.alignment
        
        # Copy merged cells
        for merged_range in existing_sheet.merged_cells.ranges:
            protocol_sheet.merge_cells(str(merged_range))
        
        # Copy column dimensions
        for col_letter, col_dimension in existing_sheet.column_dimensions.items():
            protocol_sheet.column_dimensions[col_letter].width = col_dimension.width
            
        # Copy row dimensions
        for row_num, row_dimension in existing_sheet.row_dimensions.items():
            protocol_sheet.row_dimensions[row_num].height = row_dimension.height
            
        existing_wb.close()
        
    except Exception as e:
        print(f"Error loading existing protocol template: {e}")
        # Create a simple protocol sheet as fallback
        protocol_sheet.cell(row=1, column=1, value="OTIS UNIFIED PROTOCOL TEMPLATE")
        protocol_sheet.cell(row=2, column=1, value="This sheet contains the protocol layout for data insertion")
    
    # Save unified template
    output_path = "/home/runner/workspace/EGYESITETT-OTIS-TEMPLATE.xlsx"
    wb.save(output_path)
    wb.close()
    
    print(f"✅ Unified template created successfully: {output_path}")
    print("📊 Template contains:")
    print("   - questions sheet: All question definitions (21 original + 5 measurement)")
    print("   - protocol sheet: Complete OTIS protocol layout for data insertion")
    print("   - Measurement data block integrated with traditional questions")
    
    return output_path

if __name__ == "__main__":
    create_unified_template()