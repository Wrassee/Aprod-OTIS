#!/usr/bin/env python3
"""
Comprehensive OTIS APROD Template Creator
Creates a complete Excel template with all supported question types and features
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_comprehensive_template():
    # Create comprehensive question data with all supported types and features
    questions_data = [
        # Basic text questions
        {
            'ID': '1',
            'Title_Hun': 'Szerelő neve',
            'Title_De': 'Name des Monteurs',
            'Type': 'text',
            'Cél/Placeholder': 'Teljes név megadása',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Általános adatok',
            'Order': 1,
            'Kell': 'igen',
            'Leírás/Placeholder': 'A szerelő teljes neve',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Allgemeine Daten',
            'MultiCell': 'nem'
        },
        {
            'ID': '2', 
            'Title_Hun': 'Átvevő neve',
            'Title_De': 'Name des Übernehmers',
            'Type': 'text',
            'Cél/Placeholder': 'F12',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Általános adatok',
            'Order': 2,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Az átvevő teljes neve',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Allgemeine Daten',
            'MultiCell': 'nem'
        },
        
        # Number/measurement questions
        {
            'ID': 'm1',
            'Title_Hun': 'Kabinendach - Schachtkopf távolság',
            'Title_De': 'Abstand Kabinendach - Schachtkopf',
            'Type': 'number',
            'Cél/Placeholder': 'G15',
            'Unit': 'mm',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '1000',
            'max_value': '3000',
            'Blokk neve HU': 'Mérések',
            'Order': 1,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Távolság milliméterben',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Messungen',
            'MultiCell': 'nem'
        },
        {
            'ID': 'm2',
            'Title_Hun': 'Effektív biztonsági távolság A',
            'Title_De': 'Effektiver Sicherheitsabstand A',
            'Type': 'measurement',
            'Cél/Placeholder': 'G20',
            'Unit': 'mm',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '800',
            'max_value': '2500',
            'Blokk neve HU': 'Mérések',
            'Order': 2,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Biztonsági távolság mérése',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Messungen',
            'MultiCell': 'nem'
        },
        
        # Yes/No questions
        {
            'ID': '10',
            'Title_Hun': 'Kabinentür beépítve?',
            'Title_De': 'Kabinentür eingebaut?',
            'Type': 'yes_no',
            'Cél/Placeholder': 'H25',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Ellenőrzések',
            'Order': 1,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Igen/Nem választás',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Überprüfungen',
            'MultiCell': 'nem'
        },
        {
            'ID': '11',
            'Title_Hun': 'VF Inverter beépített?',
            'Title_De': 'VF Inverter eingebaut?',
            'Type': 'yes_no_na',
            'Cél/Placeholder': 'H30',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Ellenőrzések',
            'Order': 2,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Igen/Nem/N.A. választás',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Überprüfungen',
            'MultiCell': 'igen'
        },
        
        # Dropdown/select questions
        {
            'ID': '15',
            'Title_Hun': 'Kontroller típusa',
            'Title_De': 'Controller-Typ',
            'Type': 'select',
            'Cél/Placeholder': 'I35|OTIS Gen2|OTIS Compass|OTIS MRL|Egyéb',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Berendezések',
            'Order': 1,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Válasszon a listából',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Geräte',
            'MultiCell': 'nem'
        },
        
        # Calculated questions
        {
            'ID': 'calc1',
            'Title_Hun': 'Teljes lift magasság',
            'Title_De': 'Gesamthöhe des Aufzugs',
            'Type': 'calculated',
            'Cél/Placeholder': 'J40',
            'Unit': 'mm',
            'calculation_formula': 'A + B + C',
            'calculation_inputs': 'm1,m2,m3',
            'min_value': '2000',
            'max_value': '50000',
            'Blokk neve HU': 'Számítások',
            'Order': 1,
            'Kell': 'nem',
            'Leírás/Placeholder': 'Automatikusan számított érték',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Berechnungen',
            'MultiCell': 'nem'
        },
        
        # Additional measurement for niedervolt
        {
            'ID': 'm3',
            'Title_Hun': 'Nachlaufweg mérés',
            'Title_De': 'Nachlaufweg Messung',
            'Type': 'measurement',
            'Cél/Placeholder': 'K45',
            'Unit': 'mm',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '100',
            'max_value': '1000',
            'Blokk neve HU': 'Mérések',
            'Order': 3,
            'Kell': 'igen',
            'Leírás/Placeholder': 'Nachlaufweg távolság',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Messungen',
            'MultiCell': 'nem'
        },
        
        # Date question
        {
            'ID': 'date1',
            'Title_Hun': 'Átadás dátuma',
            'Title_De': 'Übergabedatum',
            'Type': 'date',
            'Cél/Placeholder': 'L50',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Dokumentáció',
            'Order': 1,
            'Kell': 'igen',
            'Leírás/Placeholder': 'ÉÉÉÉ-HH-NN formátum',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Dokumentation',
            'MultiCell': 'nem'
        },
        
        # Textarea question
        {
            'ID': 'notes1',
            'Title_Hun': 'Megjegyzések',
            'Title_De': 'Anmerkungen',
            'Type': 'textarea',
            'Cél/Placeholder': 'M55',
            'Unit': '',
            'calculation_formula': '',
            'calculation_inputs': '',
            'min_value': '',
            'max_value': '',
            'Blokk neve HU': 'Dokumentáció',
            'Order': 2,
            'Kell': 'nem',
            'Leírás/Placeholder': 'Többsoros szöveg',
            'Munkalap neve': 'Protocol',
            'Blokk neve DE': 'Dokumentation',
            'MultiCell': 'nem'
        }
    ]
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # 1. Questions sheet
    questions_ws = wb.create_sheet("questions")
    
    # Create DataFrame and add to worksheet
    df = pd.DataFrame(questions_data)
    
    # Add header with styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for r in dataframe_to_rows(df, index=False, header=True):
        questions_ws.append(r)
    
    # Style header row
    for cell in questions_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in questions_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        questions_ws.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Niedervolt sheet for measurements
    niedervolt_ws = wb.create_sheet("niedervolt")
    
    # Add niedervolt devices data
    niedervolt_data = [
        ['Device_ID', 'Device_Name', 'Measurement_Type', 'Unit', 'Min_Value', 'Max_Value'],
        ['niv_1', 'FI kapcsoló 30mA', 'FI_Test', 'mA', '10', '30'],
        ['niv_2', 'FI kapcsoló 300mA', 'FI_Test', 'mA', '100', '300'],
        ['niv_3', 'Szigetelési ellenállás', 'Isolation', 'MΩ', '0.5', '1000'],
        ['niv_4', 'Földelési ellenállás', 'Earth', 'Ω', '0.1', '100'],
        ['niv_5', 'Vezeték folytonosság', 'Continuity', 'Ω', '0.01', '10']
    ]
    
    for row in niedervolt_data:
        niedervolt_ws.append(row)
    
    # Style niedervolt header
    for cell in niedervolt_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths for niedervolt
    for column in niedervolt_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        niedervolt_ws.column_dimensions[column_letter].width = adjusted_width
    
    # 3. Protocol template sheet (for unified templates)
    protocol_ws = wb.create_sheet("Protocol")
    
    # Add some sample cells that questions will reference
    protocol_ws['F9'] = 'Szerelő neve:'
    protocol_ws['F12'] = 'Átvevő neve:'
    protocol_ws['G15'] = 'Kabinendach-Schachtkopf:'
    protocol_ws['G20'] = 'Biztonsági távolság A:'
    protocol_ws['H25'] = 'Kabinentür:'
    protocol_ws['H30'] = 'VF Inverter:'
    protocol_ws['I35'] = 'Kontroller:'
    protocol_ws['J40'] = 'Teljes magasság:'
    protocol_ws['K45'] = 'Nachlaufweg:'
    protocol_ws['L50'] = 'Átadás dátuma:'
    protocol_ws['M55'] = 'Megjegyzések:'
    
    # Add some styling to protocol sheet
    for cell in ['F9', 'F12', 'G15', 'G20', 'H25', 'H30', 'I35', 'J40', 'K45', 'L50', 'M55']:
        protocol_ws[cell].font = Font(bold=True)
        protocol_ws[cell].fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # Save the file
    filename = 'COMPREHENSIVE-TEMPLATE-FULL.xlsx'
    wb.save(filename)
    print(f"✅ Comprehensive template created: {filename}")
    
    return filename

if __name__ == "__main__":
    create_comprehensive_template()