#!/usr/bin/env python3
"""
OTIS Egyesített Template Példa
Ez a script létrehoz egy példa Excel template-et az új szintaxis bemutatásához
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_example_template():
    """Létrehoz egy példa template-et az új formátumban"""
    
    # === QUESTIONS SHEET PÉLDA ADATOK ===
    questions_data = [
        # Header sor
        ["question_id", "title_hu", "title_de", "type", "cell_reference", "unit", "calculation_formula", "calculation_inputs", "min_value", "max_value", "group_name", "group_order", "required", "placeholder"],
        
        # ÁLTALÁNOS ADATOK CSOPORT
        ["1", "Átvevő neve", "Name des Abnehmers", "text", "F9", "", "", "", "", "", "Általános adatok", "1", "true", "Teljes név megadása"],
        ["2", "Szerelő neve", "Name des Monteurs", "text", "Q9", "", "", "", "", "", "Általános adatok", "2", "true", "Szerelő teljes neve"],
        ["3", "Irányítószám", "Postleitzahl", "number", "G13", "", "", "", "1000", "9999", "Általános adatok", "3", "true", "pl. 1051"],
        ["4", "Város", "Stadt", "text", "O13", "", "", "", "", "", "Általános adatok", "4", "true", "Város megnevezése"],
        ["5", "Utca", "Straße", "text", "G14", "", "", "", "", "", "Általános adatok", "5", "true", "Utca és házszám"],
        ["6", "Házszám", "Hausnummer", "number", "N14", "", "", "", "1", "999", "Általános adatok", "6", "true", "Házszám"],
        ["7", "OTIS Lift azonosító", "OTIS Aufzug-ID", "text", "O16", "", "", "", "", "", "Általános adatok", "7", "true", "pl. 1CV89"],
        ["8", "Projekt azonosító", "Projekt-ID", "text", "O17", "", "", "", "", "", "Általános adatok", "8", "true", "Projekt kód"],
        ["9", "Kirendeltség", "Außenstelle", "text", "O19", "", "", "", "", "", "Általános adatok", "9", "true", "Kirendeltség neve"],
        
        # GÉPHÁZ CSOPORT
        ["10", "Gépház típus", "Maschinenraumtyp", "yes_no_na", "A68,B68,C68", "", "", "", "", "", "Gépház", "1", "true", "Igen/Nem/NA"],
        ["11", "Gépház elhelyezés", "Maschinenraum Position", "yes_no_na", "A75;A76;A77,B75;B76;B77,C75;C76;C77", "", "", "", "", "", "Gépház", "2", "true", "Többsoros választás"],
        
        # MODERNIZÁCIÓ CSOPORT 
        ["12", "Kabin módosítva", "Kabine modifiziert", "true_false", "Q25", "", "", "", "", "", "Modernizációban érintett", "1", "true", "Igaz/Hamis"],
        ["13", "Ajtó cserélve", "Tür ausgetauscht", "true_false", "Q26", "", "", "", "", "", "Modernizációban érintett", "2", "true", "Igaz/Hamis"],
        ["14", "Motor frissítve", "Motor aktualisiert", "true_false", "Q27", "", "", "", "", "", "Modernizációban érintett", "3", "true", "Igaz/Hamis"],
        ["15", "Vezérlés cserélve", "Steuerung ersetzt", "true_false", "Q28", "", "", "", "", "", "Modernizációban érintett", "4", "true", "Igaz/Hamis"],
        ["16", "Kábelek frissítve", "Kabel erneuert", "true_false", "Q29", "", "", "", "", "", "Modernizációban érintett", "5", "true", "Igaz/Hamis"],
        
        # MÉRÉSI ADATOK CSOPORT ✨ ÚJ BLOKK
        ["m1", "Távolság kabintető és aknatető között", "Abstand Kabinendach zu Schachtkopf", "measurement", "I278", "mm", "", "", "500", "3000", "Mérési adatok", "1", "true", "Mérés mm-ben"],
        ["m2", "Kabintető legmagasabb pont távolsága", "Höchster Punkt Kabinendach Abstand", "measurement", "N278", "mm", "", "", "400", "2500", "Mérési adatok", "2", "true", "Mérés mm-ben"],
        ["m3", "Aknapadló-ellensúly puffer távolság", "Schachtboden-Gegengewicht Puffer", "measurement", "I280", "mm", "", "", "200", "1500", "Mérési adatok", "3", "true", "Mérés mm-ben"],
        ["m4", "Effektív biztonsági távolság A", "Effektiver Sicherheitsabstand A", "calculated", "I283", "mm", "m1 - m3", "m1,m3", "700", "9000", "Mérési adatok", "4", "true", "Automatikusan számolt"],
        ["m5", "Effektív biztonsági távolság B", "Effektiver Sicherheitsabstand B", "calculated", "N283", "mm", "m2 - m3", "m2,m3", "400", "9000", "Mérési adatok", "5", "true", "Automatikusan számolt"],
        
        # TOVÁBBI PÉLDA KÉRDÉSEK
        ["20", "Lift típus", "Aufzugstyp", "text", "A68", "", "", "", "", "", "Műszaki adatok", "1", "true", "pl. személylift"],
        ["21", "Terhelhetőség", "Tragfähigkeit", "number", "B68", "kg", "", "", "100", "5000", "Műszaki adatok", "2", "true", "kg-ban"],
        ["22", "Sebesség", "Geschwindigkeit", "number", "C68", "m/s", "", "", "0.5", "4.0", "Műszaki adatok", "3", "true", "m/s-ban"],
        ["23", "Megállók száma", "Anzahl Haltestellen", "number", "D68", "db", "", "", "2", "50", "Műszaki adatok", "4", "true", "szintek száma"],
    ]
    
    # Excel workbook létrehozása
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Alapértelmezett sheet törlése
    
    # === QUESTIONS SHEET ===
    questions_sheet = wb.create_sheet("questions")
    
    # Adatok írása
    for row_idx, row_data in enumerate(questions_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = questions_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Header styling
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Oszlopszélességek beállítása
    col_widths = [12, 35, 35, 12, 15, 8, 20, 15, 10, 10, 20, 12, 10, 25]
    for i, width in enumerate(col_widths, 1):
        questions_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # === PROTOCOL SHEET ===
    protocol_sheet = wb.create_sheet("protocol")
    
    # Protocol sheet tartalom (egyszerűsített példa)
    protocol_data = [
        ["OTIS EGYESÍTETT PROTOKOLL TEMPLATE - PÉLDA"],
        [""],
        ["Ez a lap tartalmazza az OTIS protokoll sablont."],
        ["A kérdésekből származó adatok konkrét cellákba kerülnek."],
        [""],
        ["PÉLDA CELLÁK A KÉRDÉSEKHEZ:"],
        [""],
        ["F9: Átvevő neve (question_id: 1)"],
        ["Q9: Szerelő neve (question_id: 2)"],
        ["G13: Irányítószám (question_id: 3)"],
        ["I278: m1 mérési érték (question_id: m1)"],
        ["I283: m4 számított érték (question_id: m4)"],
        [""],
        ["KÉRDÉS TÍPUSOK MAGYARÁZAT:"],
        [""],
        ["text: Szöveges bevitel"],
        ["number: Numerikus érték (min/max validációval)"],
        ["yes_no_na: Igen/Nem/Nem alkalmazható"],
        ["true_false: Igaz/Hamis (X/- jellel az Excelben)"],
        ["measurement: Mérési adat (egységgel, tartománnyal)"],
        ["calculated: Számított érték (képlet alapján)"],
    ]
    
    for row_idx, row_data in enumerate(protocol_data, 1):
        cell = protocol_sheet.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "PÉLDA CELLÁK" in str(row_data[0]) or "KÉRDÉS TÍPUSOK" in str(row_data[0]):
            cell.font = Font(bold=True, color="366092")
    
    # Excel mentése
    output_path = "/home/runner/workspace/OTIS-TEMPLATE-PELDA.xlsx"
    wb.save(output_path)
    wb.close()
    
    print(f"✅ Példa template létrehozva: {output_path}")
    return output_path

if __name__ == "__main__":
    create_example_template()